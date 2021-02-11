from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox as mb
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib



file = pathlib.Path('Student_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"] = "Full Name"
    sheet["B1"] = "Adm. No"
    sheet["C1"] = "Class"
    sheet["D1"] = "Phone No"
    sheet["E1"] = "Residence"
    sheet["F1"] = "DOB"
    sheet["G1"] = "Faculty"

    file.save('Student_data.xlsx')


class Registration:

    def __init__(self,reg):
        # Labels for the Entry form
        self.lbl1 = Label(reg, text="Student Name:", font=("Arial Black", 8))
        self.lbl2 = Label(reg, text="Adm. Number:", font=("Arial Black", 8))
        self.lbl3 = Label(reg, text="Class:", font=("Arial Black", 8))
        self.lbl4 = Label(reg, text="Phone Number:", font=("Arial Black", 8))
        self.lbl5 = Label(reg, text="Residence:", font=("Arial Black", 8))
        self.lbl6 = Label(reg, text="Date of Birth:", font=("Arial Black", 8))
        # label for DropDown for Faculty/College/School
        self.lbl7 = Label(reg, text="Faculty / College: ", font=("Arial Black", 8)).grid(column=0,
                                                     row=5, padx=150, pady=200)

        # Combobox creation
        n = StringVar()
        facultychoosen = self.lbl7 = Combobox(reg, width=27, textvariable=n)

        # Adding combobox drop down list
        facultychoosen['values'] = (' Business',
                                  ' Technology',
                                  ' Ed. & Arts',
                                  )

        facultychoosen.grid(column=1, row=5,)
        facultychoosen.current()

        # Buttons for entry form
        self.btn1 = Button(reg, text="Insert", fg='yellow', bg='black', font=("Arial Black",8), width=5, command=self.enter)
        self.btn2 = Button(reg, text="Update", fg='yellow', bg='black', font=("Arial Black",8), width=5)
        self.btn3 = Button(reg, text="Delete", fg='yellow', bg='black', font=("Arial Black",8), width=5)
        self.btn4 = Button(reg, text="Get", fg='yellow', bg='black', font=("Arial Black",8), width=5)
        self.btn5 = Button(reg, text="Clear",fg='yellow', bg='black', font=("Arial Black",8), width=5, command=self.clear)
        self.btn6 = Button(reg, text="Quit", fg='yellow', bg='black', font=("Arial Black",8), width=5, command=self.leave)
        self.btn1.place(x=30,y=50)
        self.btn2.place(x=30,y=100)
        self.btn3.place(x=30,y=150)
        self.btn4.place(x=30,y=200)
        self.btn5.place(x=30,y=250)
        self.btn6.place(x=30,y=300)
















        self.entr1 = Entry(fg='black', bg='white')
        self.entr2 = Entry(fg='black', bg='white')
        self.entr3 = Entry(fg='black', bg='white')
        self.entr4 = Entry(fg='black', bg='white')
        self.entr5 = Entry(fg='black', bg='white')
        self.entr6 = Entry(fg='black', bg='white')
        self.lbl1.place(x=150, y=50)
        self.entr1.place(x=250, y=50)
        self.lbl2.place(x=450, y=50)
        self.entr2.place(x=550, y=50)
        self.lbl3.place(x=150, y=100)
        self.entr3.place(x=250, y=100)
        self.lbl4.place(x=450, y=100)
        self.entr4.place(x=550, y=100)
        self.lbl5.place(x=150, y=150)
        self.entr5.place(x=250, y=150)
        self.lbl6.place(x=450, y=150)
        self.entr6.place(x=550, y=150)



# Function for the quit button to exit the program
    def leave(self):
        if mb.askyesno('Exit Program', 'Do you want to quit?'):
            window.destroy()
        else:
            mb.showinfo('Aborted', 'Quit has been cancelled')

    def enter(self):
        a=self.entr1.get()
        b=self.entr2.get()
        c=self.entr3.get()
        d=self.entr4.get()
        e=self.entr5.get()
        f=self.entr6.get()
        g=self.lbl7.get()

        print(a)
        print(b)
        print(c)
        print(d)
        print(e)
        print(f)
        print(g)


        file=openpyxl.load_workbook("Student_data.xlsx")
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=a)
        sheet.cell(column=2,row=sheet.max_row,value=b)
        sheet.cell(column=3,row=sheet.max_row,value=c)
        sheet.cell(column=4,row=sheet.max_row,value=d)
        sheet.cell(column=5,row=sheet.max_row,value=e)
        sheet.cell(column=6,row=sheet.max_row,value=f)
        sheet.cell(column=7,row=sheet.max_row,value=g)

        self.entr1.delete(0, END)
        self.entr2.delete(0, END)
        self.entr3.delete(0, END)
        self.entr4.delete(0, END)
        self.entr5.delete(0, END)
        self.entr6.delete(0, END)
        self.lbl7.set("")


        file.save("Student_data.xlsx")

    def clear(self):
        self.entr1.delete(0, END)
        self.entr2.delete(0, END)
        self.entr3.delete(0, END)
        self.entr4.delete(0, END)
        self.entr5.delete(0, END)
        self.entr6.delete(0, END)
        self.lbl7.set("")


































window=Tk()
window.title("Student Registration")
mywin=Registration(window)
lbl=Label(window, text="STUDENT REGISTRATION", fg='black', font=("Arial Black", 16))
lbl.place(x=0,y=0)
window.geometry("900x700+10+20")
window.mainloop()

